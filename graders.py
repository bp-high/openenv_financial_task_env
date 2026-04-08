"""Grading functions for the Financial Task Environment.

Two grading modes:
  1. QA tasks  — compare agent text answer against reference text
                 (numeric extraction + keyword matching)
  2. MODIFY tasks — compare agent-produced xlsx against reference xlsx
                    (cell-level comparison with tolerance)
"""

from __future__ import annotations

import re
import traceback
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

import openpyxl
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Numeric helpers
# ---------------------------------------------------------------------------

def _extract_numbers(text: str) -> List[float]:
    """Extract all numeric values from text, handling commas, $, %."""
    cleaned = text.replace("$", "").replace("€", "").replace("£", "")
    pattern = r"-?\d{1,3}(?:,\d{3})*(?:\.\d+)?|-?\d+(?:\.\d+)?"
    raw = re.findall(pattern, cleaned)
    results: List[float] = []
    for r in raw:
        try:
            results.append(float(r.replace(",", "")))
        except ValueError:
            continue
    return results


def _number_close(actual: float, expected: float, rel_tol: float = 0.05) -> bool:
    if expected == 0:
        return abs(actual) < 1e-6
    return abs(actual - expected) / abs(expected) <= rel_tol


def _best_number_match(numbers: List[float], target: float, rel_tol: float = 0.05) -> bool:
    return any(_number_close(n, target, rel_tol) for n in numbers)


# ---------------------------------------------------------------------------
# QA grading (text answer)
# ---------------------------------------------------------------------------

def grade_qa(answer: str, reference_answer: str) -> float:
    """Grade a text answer against a reference.  Returns 0.0–1.0."""
    if not answer.strip():
        return 0.0

    ref_nums = _extract_numbers(reference_answer)
    ans_nums = _extract_numbers(answer)

    if ref_nums:
        # Numeric comparison: what fraction of reference numbers appear?
        matched = sum(1 for r in ref_nums if _best_number_match(ans_nums, r))
        num_score = matched / len(ref_nums)
    else:
        num_score = 0.0

    # Keyword overlap
    ref_words = set(re.findall(r"[a-zA-Z]{3,}", reference_answer.lower()))
    ans_words = set(re.findall(r"[a-zA-Z]{3,}", answer.lower()))
    if ref_words:
        kw_score = len(ref_words & ans_words) / len(ref_words)
    else:
        kw_score = 0.0

    # Weighted combination (numbers matter more for financial tasks)
    if ref_nums:
        # If all numbers match perfectly, give full score
        if num_score >= 1.0:
            return 1.0
        return round(min(0.8 * num_score + 0.2 * kw_score, 1.0), 4)
    else:
        return round(kw_score, 4)


# ---------------------------------------------------------------------------
# MODIFY grading (xlsx comparison)
# ---------------------------------------------------------------------------

def _load_wb_values(path: str):
    """Load workbook in data_only mode, return dict of {(sheet, row, col): value}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    cells = {}
    sheets = set()
    for name in wb.sheetnames:
        sheets.add(name)
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cells[(name, cell.row, cell.column)] = cell.value
    wb.close()
    return cells, sheets


def grade_xlsx(output_path: str, reference_path: str) -> float:
    """Compare agent output xlsx with reference xlsx.  Returns 0.0–1.0.

    Scoring breakdown:
      - 30%  sheet-level: does the output have all reference sheets?
      - 70%  cell-level:  fraction of reference cells matched (with tolerance for numbers)
    """
    try:
        ref_cells, ref_sheets = _load_wb_values(reference_path)
        out_cells, out_sheets = _load_wb_values(output_path)
    except Exception:
        return 0.0

    # --- Sheet score (30%) ---
    if ref_sheets:
        sheet_score = len(ref_sheets & out_sheets) / len(ref_sheets)
    else:
        sheet_score = 1.0

    # --- Cell score (70%) ---
    if not ref_cells:
        return round(0.3 * sheet_score + 0.7 * 1.0, 4)

    matched = 0
    total = len(ref_cells)

    for key, ref_val in ref_cells.items():
        out_val = out_cells.get(key)
        if out_val is None:
            continue
        if ref_val == out_val:
            matched += 1
            continue
        # Numeric tolerance
        try:
            rv = float(ref_val)
            ov = float(out_val)
            if _number_close(ov, rv, rel_tol=0.02):
                matched += 1
                continue
        except (ValueError, TypeError):
            pass
        # String comparison (case-insensitive, whitespace-normalized)
        try:
            if str(ref_val).strip().lower() == str(out_val).strip().lower():
                matched += 1
        except Exception:
            pass

    cell_score = matched / total if total > 0 else 1.0

    return round(0.3 * sheet_score + 0.7 * cell_score, 4)


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

def _clamp_score(score: float) -> float:
    """Clamp score to open interval (0.001, 0.999) — evaluator rejects exact 0.0 and 1.0."""
    return max(0.001, min(0.999, score))


def grade_task(task: Dict[str, Any], answer: str = "", output_path: str = "") -> float:
    """Grade a task.  Returns score in (0.001, 0.999).

    For QA tasks:    uses *answer* (text) vs task["reference_answer"].
    For MODIFY tasks: uses *output_path* (xlsx) vs task["reference_file"].
    """
    task_type = task.get("task_type", "QA")

    if task_type == "QA":
        ref = task.get("reference_answer", "")
        return _clamp_score(grade_qa(answer, ref))
    elif task_type == "MODIFY":
        ref_path = task.get("reference_file", "")
        if not output_path or not ref_path:
            return 0.001
        if not Path(output_path).exists() or not Path(ref_path).exists():
            return 0.001
        return _clamp_score(grade_xlsx(output_path, ref_path))
    else:
        return 0.001
