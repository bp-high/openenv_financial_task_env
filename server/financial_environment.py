"""Financial Task Environment — core environment logic.

A code-execution environment where the agent writes Python code (using openpyxl)
to read, analyze, and modify real Excel workbooks from enterprise finance workflows.

For QA tasks: the agent reads the xlsx and submits a text answer.
For MODIFY tasks: the agent writes code that modifies the xlsx, then the result
is compared cell-by-cell against a reference workbook.
"""

from __future__ import annotations

import io
import openpyxl
import os
import shutil
import subprocess
import sys
import tempfile
import traceback
from pathlib import Path
from typing import Any, Optional
from uuid import uuid4

from openenv.core.env_server.interfaces import Environment
from openenv.core.env_server.types import Observation, State

from models import FinancialAction, FinancialObservation
from tasks import TASKS, TASK_IDS, get_task
from graders import grade_task
from rewards import RewardTracker


# OpenEnv concurrency support — each WebSocket session gets its own
# FinancialEnvironment instance.  Per-instance state (workdir, gold stash,
# code-step counter) is fully session-isolated; gold-stash naming uses
# uuid4-randomized tmpdirs that won't collide.
SUPPORTS_CONCURRENT_SESSIONS: bool = True


class FinancialEnvironment(Environment):
    """OpenEnv environment for financial spreadsheet tasks with code execution.

    Episode flow
    ────────────
    1. ``reset(task_id="task_1")`` → observation with task info + xlsx summary.
    2. ``step(action_type="code", content="import openpyxl; ...")`` → execute code, get stdout.
    3. ``step(action_type="submit", content="answer text")`` → grade and end episode.
       *or* for MODIFY tasks:
       ``step(action_type="submit_file", content="<path>")`` → grade xlsx and end.

    The episode also ends when *max_steps* is reached.
    """

    MAX_STEPS = 15

    def __init__(self) -> None:
        super().__init__()
        self._state = State(episode_id=str(uuid4()), step_count=0)
        self._current_task: dict[str, Any] | None = None
        self._done = False
        self._cumulative_reward = 0.0
        self._workdir: str | None = None
        self._gold_stash_dir: str | None = None  # per-episode stash for gold files
        self._gold_originals: list[tuple[str, str]] = []  # (original_path, stashed_path) for restore
        self._reward_tracker: RewardTracker | None = None
        # Progress signal (distance-to-gold) is on by default; set
        # FINANCIAL_ENV_PROGRESS=0 to disable for clean eval.
        self._progress_enabled = os.environ.get("FINANCIAL_ENV_PROGRESS", "1") == "1"
        # Minimum number of code steps before a submit is accepted.  Default 1
        # — kills the "submit at step 1 with source unchanged" exploit at the
        # env layer (post-grading defense alone wasn't enough — Kimi still
        # tried it even when the grader scored 0.001).  Set to 0 to disable.
        self._min_code_steps_before_submit = int(
            os.environ.get("FINANCIAL_ENV_MIN_CODE_STEPS", "1")
        )
        self._code_steps_taken = 0

    # ------------------------------------------------------------------
    # reset
    # ------------------------------------------------------------------
    def reset(
        self,
        seed: Optional[int] = None,
        episode_id: Optional[str] = None,
        **kwargs: Any,
    ) -> FinancialObservation:
        # Clean up any leftover state from a prior episode.  In particular this
        # restores stashed gold files to their data/ locations — without this,
        # back-to-back resets would leak gold permanently into /tmp.
        if self._workdir is not None or self._gold_stash_dir is not None:
            self.close()

        task_id: str = kwargs.get("task_id", "task_1")
        # IMPORTANT: copy the global task dict — we'll mutate per-episode fields
        # (reference_file, evaluator.checks[*].expected_files) to point at a
        # stashed location the agent can't easily glob for.  Mutating the global
        # would break subsequent episodes.
        self._current_task = dict(get_task(task_id))
        self._state = State(
            episode_id=episode_id or str(uuid4()),
            step_count=0,
        )
        self._done = False
        self._cumulative_reward = 0.0
        self._code_steps_taken = 0

        # Create a working directory and copy the source xlsx into it
        self._workdir = tempfile.mkdtemp(prefix=f"financial_env_{task_id}_")
        src = self._current_task.get("source_file", "")
        if src and Path(src).exists():
            shutil.copy2(src, self._workdir)
            work_file = str(Path(self._workdir) / Path(src).name)
        else:
            work_file = ""

        # Stash gold files into a private tmpdir.  This is best-effort sandboxing:
        # the dir name is uuid-randomized and lives outside the repo data tree, so
        # naive `glob('/app/env/data/**/*_ref*')` won't find it.  A determined
        # agent that scans /tmp can still find it — for that we'd need bwrap or
        # seccomp.  This kills the most common reward-hacking vector.
        self._gold_stash_dir = tempfile.mkdtemp(prefix="oe_gold_")
        self._stash_gold_files(self._current_task, Path(self._gold_stash_dir))

        # Stand up the per-episode reward tracker — *after* gold stashing, so the
        # tracker's progress-distance fn reads the stashed (hidden) gold path.
        family = self._current_task.get("family", "xlsx")
        gold = self._current_task.get("reference_file", "")
        if work_file and family in ("xlsx", "pptx", "docx"):
            try:
                # Per-task evaluator callable for the new 6th reward signal.
                # Only docx tasks currently have one; xlsx/pptx pass None.
                task_eval = self._make_task_evaluator() if family == "docx" else None
                self._reward_tracker = RewardTracker(
                    family=family,
                    working_file=work_file,
                    gold_file=gold or None,
                    enable_progress=self._progress_enabled,
                    task_evaluator=task_eval,
                )
            except Exception:
                self._reward_tracker = None
        else:
            self._reward_tracker = None

        # Generate an xlsx summary to include in the observation
        xlsx_summary = self._summarize_xlsx(work_file) if work_file else "No source file."

        task = self._current_task
        task_info = (
            f"Task: {task['title']}\n"
            f"Difficulty: {task['difficulty']}\n"
            f"Type: {task['task_type']} ({task['category']})\n\n"
            f"Instruction:\n{task['instruction']}\n"
        )
        if task.get("constraints"):
            task_info += f"\nConstraints:\n{task['constraints']}\n"
        task_info += (
            f"\nSource file: {work_file}\n"
            f"\nSpreadsheet Summary:\n{xlsx_summary}\n\n"
            "Actions:\n"
            "  action_type='code'    → Execute Python code (openpyxl available).\n"
            "                          The working file path is in the source_file field.\n"
            "  action_type='submit'  → Submit a text answer (QA tasks).\n"
            "  action_type='submit_file' → Submit a modified xlsx path (MODIFY tasks).\n"
        )

        return FinancialObservation(
            done=False,
            reward=0.0,
            task_id=task["id"],
            task_description=task_info,
            financial_data=xlsx_summary,
            difficulty=task["difficulty"],
            task_type=task["task_type"],
            feedback="Environment reset. Read the spreadsheet and task instructions carefully.",
            current_step=0,
            max_steps=self.MAX_STEPS,
            available_tasks=",".join(TASK_IDS),
            source_file=work_file,
        )

    # ------------------------------------------------------------------
    # step
    # ------------------------------------------------------------------
    def step(
        self,
        action: FinancialAction,
        timeout_s: Optional[float] = None,
        **kwargs: Any,
    ) -> FinancialObservation:
        self._state.step_count += 1

        if self._current_task is None:
            return self._obs(feedback="No task loaded. Call reset() first.", reward=0.001, done=True)

        if self._done:
            return self._obs(feedback="Episode already finished. Call reset().", reward=0.001, done=True)

        action_type = action.action_type.strip().lower()

        if action_type == "code":
            return self._handle_code(action.content)
        elif action_type == "submit":
            return self._handle_submit_text(action.content)
        elif action_type == "submit_file":
            return self._handle_submit_file(action.content)
        else:
            return self._obs(
                feedback=f"Unknown action_type '{action.action_type}'. Use 'code', 'submit', or 'submit_file'.",
                reward=0.001, done=False,
            )

    # ------------------------------------------------------------------
    # Gold stashing — defense against reward hacking via gold-file-read
    # ------------------------------------------------------------------
    def _stash_gold_files(self, task: dict[str, Any], stash_dir: Path) -> None:
        """Stash all gold files referenced by `task` into `stash_dir` and
        rewrite the task's gold-file paths to point at the stashed copies.

        Two modes (set via `FINANCIAL_ENV_GOLD_STASH=move|copy`, default `move`):

        * **move** (default): rent the gold file to a uuid-randomized tmp path
          for the duration of the episode; close() puts it back.  Strongest
          defense — agent globbing data/ finds nothing.  Single-tenant: two
          parallel episodes of the same task would race.

        * **copy**: each session gets its own copy, originals stay in data/.
          Concurrent-friendly (required for GRPO with N>1 generations on the
          same task).  Trades the data/-glob defense for parallelism, but
          Phase 7 byte-equality + Phase 9 early-submit gate still block
          'submit gold/source unchanged' exploits independently.

        For training Spaces, set `FINANCIAL_ENV_GOLD_STASH=copy` so concurrent
        rollouts don't fight over the same source's rename.
        """
        from secrets import token_hex

        mode = os.environ.get("FINANCIAL_ENV_GOLD_STASH", "move").lower()
        moves: list[tuple[str, str]] = []  # (original, stashed) for restore (move mode only)
        path_map: dict[str, str] = {}      # original -> stashed (dedup within session)

        def _stash(src: str, label: str) -> str:
            """Stash src into stash_dir.  Idempotent within a session: same
            src always maps to the same stashed path."""
            if src in path_map:
                return path_map[src]
            suffix = Path(src).suffix or ""
            dest = stash_dir / f"{label}_{token_hex(3)}{suffix}"

            if mode == "copy":
                # Concurrent-safe: each session gets its own private copy
                try:
                    shutil.copy2(src, dest)
                except FileNotFoundError:
                    # Source already moved by a concurrent session — the
                    # data/ original is gone, but that's OK in copy mode if
                    # another session published its stash to a known path.
                    # For correctness, just record the original path; grader
                    # will fail gracefully via the existence check upstream.
                    path_map[src] = src
                    return src
                # No restore-on-close needed in copy mode (originals untouched)
                path_map[src] = str(dest)
                return str(dest)

            # move mode (default): atomic rename, restore on close()
            try:
                Path(src).rename(dest)  # atomic on same FS
            except OSError:
                shutil.copy2(src, dest)
                try:
                    Path(src).unlink()
                except OSError:
                    pass  # best-effort
            moves.append((src, str(dest)))
            path_map[src] = str(dest)
            return str(dest)

        # 1. DOCX evaluator's per-check expected_files (multi-gold support)
        # Process this FIRST so that when we later resolve reference_file, we
        # find the already-stashed location via path_map (handles the common
        # case where reference_file == evaluator.checks[0].expected_files[0]).
        evaluator = task.get("evaluator")
        if evaluator and "checks" in evaluator:
            new_checks: list[dict[str, Any]] = []
            for c_idx, check in enumerate(evaluator.get("checks", []) or []):
                new_check = dict(check)
                new_expected: list[str] = []
                for f_idx, ef in enumerate(check.get("expected_files") or []):
                    if ef and Path(ef).exists():
                        new_expected.append(_stash(ef, f"check_{c_idx}_{f_idx}"))
                    else:
                        # Already moved earlier this episode? Resolve via map.
                        new_expected.append(path_map.get(ef, ef))
                new_check["expected_files"] = new_expected
                new_checks.append(new_check)
            task["evaluator"] = {**evaluator, "checks": new_checks}

        # 2. Top-level reference_file (xlsx grader + diff layer + tracker).
        # Reuses any prior stash from step 1.
        ref = task.get("reference_file", "")
        if ref:
            if ref in path_map:
                task["reference_file"] = path_map[ref]
            elif Path(ref).exists():
                task["reference_file"] = _stash(ref, "gold_ref")
            # else: file already gone (moved earlier?) — leave path; grader will
            # gracefully degrade since exists() check upstream catches it.

        self._gold_originals = moves  # used by close() to restore

    def _make_task_evaluator(self):
        """Return a callable f(working_file)->[0,1] running the task's evaluator
        block, or None if the family has no per-task evaluator wired up.

        Used as the 6th reward signal in RewardTracker.  Imports are lazy so
        non-docx episodes don't pay for them."""
        task = self._current_task
        if not task:
            return None
        evaluator = task.get("evaluator") or {}
        checks = evaluator.get("checks") or []
        if not checks:
            return None

        from graders.docx_metrics import run_evaluator
        conj = evaluator.get("conj", "and")
        source_file = task.get("source_file", "")

        def _eval(working_file: str) -> float:
            try:
                return float(run_evaluator(
                    conj=conj,
                    checks=checks,
                    working_file=working_file,
                    source_file=source_file,
                ))
            except Exception:
                return 0.0

        return _eval

    # ------------------------------------------------------------------
    # state property
    # ------------------------------------------------------------------
    @property
    def state(self) -> State:
        return self._state

    # ------------------------------------------------------------------
    # Code execution
    # ------------------------------------------------------------------
    def _compute_code_reward(self, code: str, succeeded: bool, stdout: str) -> tuple[float, dict]:
        """Compute step reward via the unified RewardTracker.

        Returns (total, breakdown) — breakdown is a dict of named components for
        logging.  Falls back to a minimal fixed reward when no tracker is wired.
        """
        if self._reward_tracker is None:
            return (0.02 if succeeded else 0.005), {"fallback": True}

        signals = self._reward_tracker.score_step(code=code, succeeded=succeeded, stdout=stdout)
        return signals.total, signals.to_dict()

    def _handle_code(self, code: str) -> FinancialObservation:
        """Execute Python code in a subprocess and return stdout/stderr."""
        if not self._workdir:
            return self._obs(feedback="No working directory. Call reset() first.", reward=0.001, done=False)

        succeeded = False
        stdout = ""
        stderr = ""

        try:
            result = subprocess.run(
                [sys.executable, "-c", code],
                capture_output=True,
                text=True,
                timeout=30,
                cwd=self._workdir,
                env={**os.environ, "PYTHONDONTWRITEBYTECODE": "1"},
            )
            stdout = result.stdout[:4000] if result.stdout else ""
            stderr = result.stderr[:2000] if result.stderr else ""

            if result.returncode == 0:
                succeeded = True
                feedback = f"Code executed successfully.\n\nSTDOUT:\n{stdout}"
                if stderr:
                    feedback += f"\n\nSTDERR:\n{stderr}"
            else:
                feedback = f"Code execution failed (exit code {result.returncode}).\n\nSTDERR:\n{stderr}"
                if stdout:
                    feedback += f"\n\nSTDOUT:\n{stdout}"
        except subprocess.TimeoutExpired:
            feedback = "Code execution timed out (30s limit)."
        except Exception as e:
            feedback = f"Code execution error: {e}"

        reward, breakdown = self._compute_code_reward(code, succeeded, stdout)
        self._cumulative_reward += reward
        # Count this code step regardless of success — even a failed attempt
        # shows the agent at least tried before submitting.  The submit gate
        # below uses this counter to refuse early submits.
        self._code_steps_taken += 1

        # Surface the reward decomposition in feedback — useful for debugging
        # and for inspecting what the shaper credited each step.
        if breakdown and "fallback" not in breakdown:
            parts = ", ".join(f"{k}={v:.3f}" for k, v in breakdown.items() if k != "total")
            feedback += f"\n\nReward: total={breakdown['total']:.3f} ({parts})"

        at_limit = self._state.step_count >= self.MAX_STEPS
        if at_limit:
            self._done = True
            feedback += "\n\n⚠ Maximum steps reached — episode ending."

        return self._obs(feedback=feedback, reward=reward, done=at_limit)

    # ------------------------------------------------------------------
    # Submit handlers
    # ------------------------------------------------------------------
    def _early_submit_rejected(self) -> FinancialObservation | None:
        """If the agent hasn't taken enough code steps yet, reject the submit
        with a small penalty AND keep the episode open.  Returns the rejection
        observation, or None if the submit is allowed to proceed.

        Why not end the episode?  Ending makes a single bad attempt costly.
        Keeping it open lets the agent recover within its remaining budget,
        and the cost is just one wasted step — exactly the right shape of
        feedback for an RL agent learning the task.
        """
        if self._code_steps_taken >= self._min_code_steps_before_submit:
            return None
        n = self._min_code_steps_before_submit
        feedback = (
            f"❌ Submit rejected: you must execute at least {n} code step"
            f"{'' if n == 1 else 's'} before submitting (you've taken "
            f"{self._code_steps_taken} so far). "
            "Use action_type='code' to read or modify the file first, then "
            "submit. This rejection counts as one of your "
            f"{self.MAX_STEPS} steps."
        )
        # Small penalty to discourage the pattern, but episode stays open so
        # the agent has remaining steps to do real work.
        reward = 0.001
        self._cumulative_reward += reward
        at_limit = self._state.step_count >= self.MAX_STEPS
        if at_limit:
            self._done = True
            feedback += "\n\n⚠ Maximum steps reached — episode ending."
        return self._obs(feedback=feedback, reward=reward, done=at_limit)

    def _handle_submit_text(self, answer: str) -> FinancialObservation:
        """Grade a text answer (for QA tasks)."""
        early = self._early_submit_rejected()
        if early is not None:
            return early

        task = self._current_task
        assert task is not None

        score = grade_task(task, answer=answer)
        self._done = True
        self._cumulative_reward += score

        quality = "Excellent" if score >= 0.9 else "Good" if score >= 0.7 else "Partial" if score >= 0.4 else "Needs improvement"
        return self._obs(
            feedback=f"Answer graded. Score: {score:.2f}/1.00 — {quality}.\nCumulative reward: {self._cumulative_reward:.2f}",
            reward=score, done=True,
        )

    def _handle_submit_file(self, file_path: str) -> FinancialObservation:
        """Grade a modified xlsx file (for MODIFY tasks)."""
        early = self._early_submit_rejected()
        if early is not None:
            return early

        task = self._current_task
        assert task is not None

        # Resolve relative paths against workdir
        p = Path(file_path)
        if not p.is_absolute() and self._workdir:
            p = Path(self._workdir) / p

        if not p.exists():
            self._done = True
            return self._obs(
                feedback=f"File not found: {p}. Score: 0.001",
                reward=0.001, done=True,
            )

        score = grade_task(task, output_path=str(p))
        self._done = True
        self._cumulative_reward += score

        quality = "Excellent" if score >= 0.9 else "Good" if score >= 0.7 else "Partial" if score >= 0.4 else "Needs improvement"
        return self._obs(
            feedback=f"File graded. Score: {score:.2f}/1.00 — {quality}.\nCumulative reward: {self._cumulative_reward:.2f}",
            reward=score, done=True,
        )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _summarize_xlsx(self, path: str) -> str:
        """Return a text summary of an xlsx file (sheet names, dimensions, sample data)."""
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            lines = [f"Workbook: {Path(path).name}", f"Sheets: {wb.sheetnames}", ""]
            for name in wb.sheetnames[:5]:  # Limit to 5 sheets
                ws = wb[name]
                lines.append(f"--- Sheet: {name} (rows≈{ws.max_row}, cols≈{ws.max_column}) ---")
                # Show first 8 rows
                row_count = 0
                for row in ws.iter_rows(max_row=8, values_only=True):
                    vals = [str(v)[:30] if v is not None else "" for v in row[:12]]
                    lines.append("  " + " | ".join(vals))
                    row_count += 1
                if ws.max_row and ws.max_row > 8:
                    lines.append(f"  ... ({ws.max_row - 8} more rows)")
                lines.append("")
            wb.close()
            return "\n".join(lines)
        except Exception as e:
            return f"Could not read xlsx: {e}"

    def _obs(self, *, feedback: str, reward: float, done: bool) -> FinancialObservation:
        task = self._current_task or {}
        work_file = ""
        if self._workdir and task.get("source_file"):
            work_file = str(Path(self._workdir) / Path(task["source_file"]).name)
        return FinancialObservation(
            done=done,
            reward=reward,
            task_id=task.get("id", ""),
            task_description=task.get("instruction", ""),
            financial_data="",
            difficulty=task.get("difficulty", ""),
            task_type=task.get("task_type", ""),
            feedback=feedback,
            current_step=self._state.step_count,
            max_steps=self.MAX_STEPS,
            available_tasks=",".join(TASK_IDS),
            source_file=work_file,
        )

    def close(self) -> None:
        """Clean up per-episode tempdirs.  Restores gold files to their original
        locations under data/ (since reset() moved them out for sandboxing).
        """
        if self._workdir and Path(self._workdir).exists():
            shutil.rmtree(self._workdir, ignore_errors=True)
        self._workdir = None

        # Restore gold files to their original paths in data/ before deleting the
        # stash dir.  If restore fails (e.g. server crashed earlier), the file
        # stays in the stash and can be recovered manually.
        for original, stashed in self._gold_originals:
            try:
                if Path(stashed).exists() and not Path(original).exists():
                    Path(original).parent.mkdir(parents=True, exist_ok=True)
                    Path(stashed).rename(original)
            except OSError:
                try:
                    shutil.copy2(stashed, original)
                except OSError:
                    pass  # leave it in stash — manual recovery
        self._gold_originals = []

        if self._gold_stash_dir and Path(self._gold_stash_dir).exists():
            shutil.rmtree(self._gold_stash_dir, ignore_errors=True)
        self._gold_stash_dir = None
